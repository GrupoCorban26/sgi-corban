import asyncio
import os
import sys
from datetime import datetime

# Añadir el directorio raíz al path para poder importar módulos de la app
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.future import select
from sqlalchemy import and_
from app.database.db_connection import AsyncSessionLocal
from app.models.chat_message import ChatMessage

async def export_messages():
    print("Iniciando extracción de mensajes para entrenamiento de ML...")
    
    archivo_salida = f"mensajes_entrenamiento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    async with AsyncSessionLocal() as db:
        # Extraer mensajes ENTRANTES del CLIENTE que sean texto y tengan contenido
        query = select(ChatMessage).where(
            and_(
                ChatMessage.direccion == 'ENTRANTE',
                ChatMessage.remitente_tipo == 'CLIENTE',
                ChatMessage.tipo_contenido == 'text',
                ChatMessage.contenido != None,
                ChatMessage.contenido != ''
            )
        ).order_by(ChatMessage.created_at.desc()) # Los más recientes primero
        
        result = await db.execute(query)
        mensajes = result.scalars().all()
        
    if not mensajes:
        print("No se encontraron mensajes válidos para exportar.")
        return

    # Filtrar mensajes muy cortos (ej. "ok", "si") y eliminar duplicados exactos
    mensajes_unicos = set()
    datos_exportar = []
    
    for msg in mensajes:
        texto = msg.contenido.strip()
        # Ignorar mensajes de menos de 4 letras y evitar duplicados exactos
        if len(texto) > 3 and texto.lower() not in mensajes_unicos:
            mensajes_unicos.add(texto.lower())
            datos_exportar.append({
                "id": msg.id,
                "texto": texto,
                "intencion_etiquetar": "" # Columna vacía para que la llenen manualmente
            })

    # --- Crear libro Excel con formato profesional ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Mensajes ML"

    # Definir estilos
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_font = Font(name="Calibri", size=11)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    etiqueta_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarillo suave
    
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Encabezados
    headers = ["ID", "Texto del Mensaje", "Intención (Etiquetar)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Escribir datos
    for row_idx, dato in enumerate(datos_exportar, 2):
        # Columna ID
        cell_id = ws.cell(row=row_idx, column=1, value=dato["id"])
        cell_id.font = cell_font
        cell_id.alignment = Alignment(horizontal="center", vertical="top")
        cell_id.border = thin_border

        # Columna Texto
        cell_texto = ws.cell(row=row_idx, column=2, value=dato["texto"])
        cell_texto.font = cell_font
        cell_texto.alignment = cell_alignment
        cell_texto.border = thin_border

        # Columna Intención (con fondo amarillo para indicar que debe llenarse)
        cell_etiqueta = ws.cell(row=row_idx, column=3, value=dato["intencion_etiquetar"])
        cell_etiqueta.font = cell_font
        cell_etiqueta.alignment = cell_alignment
        cell_etiqueta.fill = etiqueta_fill
        cell_etiqueta.border = thin_border

    # Ajustar anchos de columna
    ws.column_dimensions["A"].width = 10   # ID
    ws.column_dimensions["B"].width = 80   # Texto del mensaje
    ws.column_dimensions["C"].width = 30   # Intención

    # Fijar la primera fila (encabezados) para scroll
    ws.freeze_panes = "A2"

    # Autofiltros para facilitar la búsqueda
    ws.auto_filter.ref = f"A1:C{len(datos_exportar) + 1}"

    # Guardar archivo
    wb.save(archivo_salida)
        
    print(f"[OK] Exito! Se han exportado {len(datos_exportar)} mensajes unicos.")
    print(f"[ARCHIVO] Guardado como: {archivo_salida}")
    print("\nPASOS SIGUIENTES:")
    print("1. Abre el archivo .xlsx directamente en Excel.")
    print("2. En la columna amarilla 'Intencion (Etiquetar)', escribe la intencion de cada mensaje.")
    print("   Ejemplos de etiquetas: COTIZACION, CARGA_LISTA, ASESORIA, QUEJA, OTRO.")
    print("3. Intenta etiquetar al menos 100-200 mensajes para empezar.")

if __name__ == "__main__":
    if hasattr(asyncio, 'WindowsSelectorEventLoopPolicy'):
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    asyncio.run(export_messages())
