"""
Script para pre-etiquetar automáticamente las intenciones de los mensajes
exportados en el Excel. Usa clasificación basada en keywords/reglas.

El usuario debe revisar y corregir las etiquetas después.
"""
import re
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# CATÁLOGO DE INTENCIONES Y REGLAS DE CLASIFICACIÓN
# ============================================================
# Orden de prioridad: las más específicas primero, las más genéricas al final.
# Cada regla es (nombre_intención, lista_de_patrones_regex)

REGLAS_INTENCION = [
    # --- Intenciones específicas del negocio ---
    ("COTIZACION", [
        r"\bcotiz", r"\bpresupuest", r"\bcuanto\s+(me\s+)?cuesta",
        r"\bcuánto\s+(me\s+)?cuesta", r"\btarifa", r"\bprecio",
        r"\bcuanto\s+cobr", r"\bcuánto\s+cobr", r"\bcuanto\s+sal",
        r"\bcuánto\s+sal", r"\bproforma", r"\bcosto",
    ]),
    ("INFO_CARGA", [
        r"\bcontenedor", r"\b\d+\s*hq\b", r"\b\d+\s*ft\b",
        r"\bfcl\b", r"\blcl\b", r"\bpallet", r"\bbulto",
        r"\bkilo", r"\btonelad", r"\bcarga\b", r"\bimportaci",
        r"\bexportaci", r"\bembarq", r"\bnaviera", r"\bflete",
        r"\barancel", r"\baduana", r"\bdespacho", r"\bdam\b",
        r"\bdua\b", r"\bbl\b", r"\bconocimiento de embarque",
        r"\bincoterm", r"\bfob\b", r"\bcif\b", r"\bcfr\b",
        r"\bexw\b", r"\bpartida\s+arancelaria", r"\bhs\s*code",
        r"\bmercader", r"\bmercancia", r"\benvio\b", r"\benvío\b",
    ]),
    ("SEGUIMIENTO", [
        r"\bseguimiento", r"\btracking", r"\brastreo",
        r"\bdonde\s+est[aá]", r"\bdónde\s+est[aá]",
        r"\bestado\s+(de|del|mi)", r"\ben\s+que\s+va",
        r"\ben\s+qué\s+va", r"\bya\s+lleg[oó]",
        r"\bcuando\s+lleg", r"\bcuándo\s+lleg",
        r"\bavance", r"\bstatus\b",
    ]),
    ("QUEJA", [
        r"\bqueja\b", r"\breclam", r"\bmal\s+servicio",
        r"\bno\s+me\s+respond", r"\bno\s+me\s+contest",
        r"\bno\s+me\s+atiend", r"\bdemora", r"\bpésimo",
        r"\bpesimo", r"\bmolest[oa]", r"\bincumpl",
        r"\binsatisf", r"\bdecepcion", r"\bdesilusion",
        r"\bnunca\s+(me\s+)?respond", r"\bnadie\s+(me\s+)?respond",
        r"\bestoy\s+esperando", r"\bsigo\s+esperando",
    ]),
    ("DATOS_CONTACTO", [
        r"\bmi\s+(correo|email|mail|número|numero|celular|teléfono|telefono)",
        r"\b[\w.-]+@[\w.-]+\.\w+",  # Detectar emails
        r"\bme\s+llamo\b", r"\bmi\s+nombre\s+es",
        r"\bmi\s+ruc\b", r"\bruc\s*\d", r"\bmi\s+empresa\s+es",
        r"\ble\s+paso\s+mis\s+datos", r"\bahí\s+le\s+envío",
        r"\bte\s+paso\s+(mi|el)\b", r"\benviarme\s+al\b",
    ]),
    ("DOCUMENTO", [
        r"\bfactura", r"\bboleta", r"\bguía\b", r"\bguia\b",
        r"\bcomprobante", r"\bcertificado", r"\bdocumento",
        r"\bcontrato", r"\bpóliza", r"\bpoliza",
        r"\bseguro\b",
    ]),
    ("SERVICIO_ESPECIFICO", [
        r"\bagenciamiento", r"\balmacen", r"\balmacén",
        r"\btransporte\b", r"\bdistribución", r"\bdistribucion",
        r"\boperador\s+logístic", r"\boperador\s+logistic",
        r"\bconsolidad", r"\bdesconsolidad",
        r"\bstorage\b", r"\bwarehouse",
    ]),

    # --- Intenciones conversacionales ---
    ("CONFIRMACION", [
        r"^s[ií]$", r"^si[\s,.!]*$", r"^sí[\s,.!]*$",
        r"^ok[,.\s!]*$", r"^okay", r"^dale[,.\s!]*$",
        r"^está\s+bien", r"^esta\s+bien", r"^de\s+acuerdo",
        r"^perfecto", r"^listo", r"^claro\b",
        r"^exacto", r"^correcto", r"^afirmativo",
        r"^por\s+supuesto", r"^así\s+es", r"^asi\s+es",
    ]),
    ("NEGACION", [
        r"^no[\s,.!]*$", r"^no\s+gracias", r"^nada\b",
        r"^ninguno", r"^para\s+nada", r"^en\s+absoluto",
        r"\bno\s+estoy\s+interesad", r"\bno\s+necesit",
        r"\bno\s+quiero\b", r"\bno,?\s+gracias",
    ]),
    ("CIERRE", [
        r"\beso\s+es\s+todo", r"\bnada\s+m[aá]s",
        r"\beso\s+ser[ií]a\s+todo", r"\beso\s+nomas",
        r"\beso\s+nomás", r"\bhasta\s+luego",
        r"\bhasta\s+pronto", r"\bnos\s+vemos",
        r"\bme\s+despido", r"\bchao\b", r"\badiós",
        r"\badios\b", r"\bbye\b",
    ]),
    ("CORTESIA", [
        r"\bgracias\b", r"\bagradezco", r"\bmuchas\s+gracias",
        r"\bse\s+lo\s+agradezco", r"\bmuy\s+amable",
        r"\bexcelente\s+servicio", r"\bbuen\s+servicio",
        r"\bgenial\b", r"\bque\s+bien\b", r"\bqué\s+bien\b",
    ]),
    ("SALUDO", [
        r"\bhola\b", r"\bbuenos?\s+d[ií]as?",
        r"\bbuenas?\s+tardes?", r"\bbuenas?\s+noches?",
        r"\bbuen\s+d[ií]a", r"\bqu[eé]\s+tal",
        r"\bsaludos\b", r"\bbienvenid",
    ]),
    ("CONSULTA_GENERAL", [
        r"\binformaci[oó]n", r"\bconsulta\b",
        r"\bquisiera\s+saber", r"\bme\s+puede[sn]?\s+(decir|indicar|informar|ayudar)",
        r"\bnecesito\s+(saber|ayuda|informaci)",
        r"\bpor\s+favor", r"\bme\s+podr[ií]a",
        r"\btengo\s+una\s+(duda|pregunta|consulta)",
        r"\bdeseo\s+(saber|conocer|informar)",
        r"\bquisiera\s+(conocer|informar)",
    ]),
    ("SOLICITUD_CONTACTO", [
        r"\bme\s+(puede[sn]?\s+)?llam", r"\bcomunic[aá]r(se|me)",
        r"\bcontact[aá]r(me|nos)", r"\bhabl(ar|e)\s+con",
        r"\bquiero\s+hablar", r"\bpasar\s+con\b",
        r"\bme\s+puede\s+atender", r"\bderivar",
        r"\basesor\b",
    ]),
]

# Intención por defecto si no matchea ninguna regla
INTENCION_DEFAULT = "OTRO"


def clasificar_mensaje(texto: str) -> str:
    """Clasifica un mensaje usando las reglas de keywords definidas."""
    texto_lower = texto.lower().strip()

    # Recorrer las reglas en orden de prioridad
    for intencion, patrones in REGLAS_INTENCION:
        for patron in patrones:
            if re.search(patron, texto_lower):
                return intencion

    return INTENCION_DEFAULT


def procesar_excel(ruta_archivo: str):
    """Lee el Excel, clasifica cada mensaje y guarda el resultado."""
    print(f"Leyendo archivo: {ruta_archivo}")

    wb = load_workbook(ruta_archivo)
    ws = wb.active

    # Verificar encabezados
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"Columnas encontradas: {headers}")

    # Buscar índice de columnas
    col_texto = None
    col_intencion = None
    for i, h in enumerate(headers, 1):
        if h and "texto" in h.lower():
            col_texto = i
        if h and "intenci" in h.lower():
            col_intencion = i

    if not col_texto or not col_intencion:
        print("ERROR: No se encontraron las columnas 'Texto' e 'Intencion'")
        return

    print(f"Columna de texto: {col_texto}, Columna de intencion: {col_intencion}")

    # Estilos para las etiquetas
    colores_intencion = {
        "SALUDO": "D5E8D4",           # Verde claro
        "COTIZACION": "DAE8FC",       # Azul claro
        "INFO_CARGA": "B4C7E7",       # Azul medio
        "CONSULTA_GENERAL": "E2EFDA", # Verde suave
        "SEGUIMIENTO": "FCE4D6",      # Naranja suave
        "QUEJA": "F8CBAD",            # Rojo suave
        "CONFIRMACION": "E2EFDA",     # Verde suave
        "NEGACION": "F2F2F2",         # Gris
        "CIERRE": "D9D2E9",           # Morado suave
        "CORTESIA": "FFF2CC",         # Amarillo suave
        "DATOS_CONTACTO": "C9DAF8",   # Azul lavanda
        "DOCUMENTO": "D0E0E3",        # Teal suave
        "SERVICIO_ESPECIFICO": "B6D7A8", # Verde medio
        "SOLICITUD_CONTACTO": "EAD1DC", # Rosa suave
        "OTRO": "F4CCCC",             # Rojo claro
    }

    font_etiqueta = Font(name="Calibri", size=11, bold=True)

    # Contadores para estadísticas
    conteo = {}
    total = 0

    for row in range(2, ws.max_row + 1):
        texto = ws.cell(row=row, column=col_texto).value
        if not texto:
            continue

        intencion = clasificar_mensaje(str(texto))
        total += 1
        conteo[intencion] = conteo.get(intencion, 0) + 1

        # Escribir la intención
        cell = ws.cell(row=row, column=col_intencion)
        cell.value = intencion
        cell.font = font_etiqueta
        cell.alignment = Alignment(horizontal="center", vertical="top")

        # Aplicar color según intención
        color = colores_intencion.get(intencion, "FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Guardar archivo (sobrescribir el mismo)
    wb.save(ruta_archivo)

    # Mostrar estadísticas
    print(f"\n{'='*50}")
    print(f"  RESUMEN DE CLASIFICACION AUTOMATICA")
    print(f"{'='*50}")
    print(f"  Total mensajes clasificados: {total}")
    print(f"{'='*50}")

    for intencion, cantidad in sorted(conteo.items(), key=lambda x: -x[1]):
        porcentaje = (cantidad / total) * 100
        barra = "#" * int(porcentaje / 2)
        print(f"  {intencion:<22} {cantidad:>4}  ({porcentaje:5.1f}%)  {barra}")

    print(f"{'='*50}")
    print(f"\n[OK] Archivo guardado: {ruta_archivo}")
    print("Abre el Excel y revisa/corrige las etiquetas.")
    print("Las celdas estan coloreadas por intencion para facilitar la revision.")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        archivo = sys.argv[1]
    else:
        archivo = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "mensajes_entrenamiento_20260515_123650.xlsx"
        )

    if not os.path.exists(archivo):
        print(f"ERROR: No se encontro el archivo: {archivo}")
        sys.exit(1)

    procesar_excel(archivo)
